import requests

# List of MITRE ATT&CK technique IDs
technique_ids = ["T1003", "T1004", "T1005"]

# Loop through each technique ID
for technique_id in technique_ids:
    # Construct URL for technique page
    url = f"https://attack.mitre.org/techniques/{technique_id}/"

    # Send request to website and get response
    response = requests.get(url)

    # Print HTML content to screen
    print(response.content)



import openpyxl
import re

# Specify the path to the Excel file
file_path = 'path/to/excel_file.xlsx'

# Load the workbook
workbook = openpyxl.load_workbook(file_path)

# Select the worksheet you want to extract data from
worksheet = workbook.active

# Define a list to hold the technique and sub-technique IDs found
technique_ids = []

# Define the regular expression pattern to match MITRE technique and sub-technique IDs
pattern = r'(T\d{4})\.(\d{1,2})'

# Loop through each cell in the worksheet
for row in worksheet.iter_rows(min_row=1, min_col=1, max_col=worksheet.max_column, values_only=True):
    for cell in row:
        if cell is not None:
            # Extract the sub-technique ID from the cell using the regex pattern
            match = re.search(pattern, cell)
            if match:
                technique_ids.append(match.group(0))

# Print the list of technique and sub-technique IDs
print(technique_ids)





import openpyxl
import re

# Specify the path to the Excel file
file_path = 'path/to/excel_file.xlsx'

# Load the workbook
workbook = openpyxl.load_workbook(file_path)

# Select the worksheet you want to extract data from
worksheet = workbook.active

# Define a set to hold the unique MITRE technique IDs found
technique_ids = set()

# Define the regular expression pattern to match MITRE technique IDs
pattern = r'T\d{4}'

# Loop through each row in column C, starting from the second row
for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
    cell = row[0]
    if cell.value is not None:  # Skip over empty cells
        # Extract all the MITRE technique IDs from the cell using the regex pattern
        matches = re.findall(pattern, cell.value)
        # Add the technique IDs to the set
        technique_ids.update(matches)

# Print the list of technique IDs
print(list(technique_ids))


